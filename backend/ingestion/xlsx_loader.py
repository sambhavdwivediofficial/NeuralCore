# ingestion/xlsx_loader.py
from __future__ import annotations

import asyncio
import io
from typing import Any

from ingestion.base_loader import BaseLoader, SourceType
from ingestion.loader_factory import register_loader


@register_loader(SourceType.XLSX)
class XlsxLoader(BaseLoader):
    source_type = SourceType.XLSX

    async def load(self, source_config: dict[str, Any]) -> list[dict[str, Any]]:
        raw_bytes = await self._read_bytes(source_config)
        return await asyncio.to_thread(self._parse_workbook, raw_bytes, source_config)

    def _parse_workbook(self, raw_bytes: bytes, source_config: dict[str, Any]) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        source_id = source_config.get("file_path")
        sheet_filter = source_config.get("sheet_names")
        row_as_document = source_config.get("row_as_document", True)
        documents: list[dict[str, Any]] = []

        for sheet_name in workbook.sheetnames:
            if sheet_filter and sheet_name not in sheet_filter:
                continue

            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue

            header = [str(value) if value is not None else f"column_{index}" for index, value in enumerate(header_row)]

            if row_as_document:
                for index, row in enumerate(rows_iter):
                    row_text = "\n".join(f"{column}: {value}" for column, value in zip(header, row) if value is not None)
                    if not row_text:
                        continue
                    documents.append(
                        self._build_document(
                            row_text,
                            metadata={
                                "source_type": self.source_type.value,
                                "file_path": source_id,
                                "sheet": sheet_name,
                                "row_index": index,
                                "columns": header,
                            },
                            source_id=f"{source_id}:{sheet_name}:{index}" if source_id else f"{sheet_name}:{index}",
                        )
                    )
            else:
                rows_text = [", ".join(f"{column}={value}" for column, value in zip(header, row)) for row in rows_iter]
                documents.append(
                    self._build_document(
                        "\n".join(rows_text),
                        metadata={"source_type": self.source_type.value, "file_path": source_id, "sheet": sheet_name, "columns": header},
                        source_id=f"{source_id}:{sheet_name}" if source_id else sheet_name,
                    )
                )

        workbook.close()
        return documents
